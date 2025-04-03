import os
import re
import subprocess
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment
import base64
import zipfile
from PIL import Image
import io
import shutil

class DocxToExcelPandocGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Docx to Excel (with Pandoc)")

        # StringVars for user inputs
        self.docx_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.preserve_equations = tk.BooleanVar(value=True)

        # Layout
        self.create_widgets()

    def create_widgets(self):
        # Row 0: Word doc
        tk.Label(self.master, text="Word Document:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(self.master, textvariable=self.docx_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.master, text="Browse", command=self.browse_docx).grid(row=0, column=2, padx=5, pady=5)

        # Row 1: Excel output
        tk.Label(self.master, text="Excel Output:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(self.master, textvariable=self.excel_path, width=50).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.master, text="Browse", command=self.browse_excel).grid(row=1, column=2, padx=5, pady=5)

        # Row 2: Equation options
        tk.Checkbutton(self.master, text="Preserve original LaTeX equations (keep $ symbols)", 
                      variable=self.preserve_equations).grid(row=2, column=1, padx=5, pady=5, sticky="w")

        # Row 3: Convert button
        tk.Button(self.master, text="Convert & Save", command=self.on_convert_click, width=20).grid(row=3, column=1, pady=15)

    def browse_docx(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Word Documents", "*.docx"), ("All Files", "*.*")]
        )
        if file_path:
            self.docx_path.set(file_path)
            # Auto-generate Excel output path
            if not self.excel_path.get():
                base_path = os.path.splitext(file_path)[0]
                self.excel_path.set(f"{base_path}.xlsx")

    def browse_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.excel_path.set(file_path)

    def on_convert_click(self):
        docx_file = self.docx_path.get().strip()
        excel_file = self.excel_path.get().strip()

        # Basic checks
        if not docx_file or not os.path.exists(docx_file):
            messagebox.showerror("Error", "Please select a valid .docx file.")
            return
        if not excel_file:
            messagebox.showerror("Error", "Please specify the Excel output file.")
            return

        # Extract images from the docx file
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                # Extract DOCX contents (it's a ZIP file)
                images_dir = os.path.join(tmpdir, "media")
                os.makedirs(images_dir, exist_ok=True)
                
                # Extract images from DOCX
                self.extract_images_from_docx(docx_file, images_dir)
                
                # Convert docx -> .tex using pandoc
                tex_path = os.path.join(tmpdir, "converted.tex")
                cmd = ["pandoc", docx_file, "-o", tex_path]
                subprocess.run(cmd, check=True)

                # Parse the generated .tex for MCQs in the desired structure
                mcq_data = self.parse_latex_for_mcqs(tex_path, images_dir)
                
        except FileNotFoundError:
            messagebox.showerror("Pandoc Error", "pandoc not found. Please install pandoc.")
            return
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Pandoc Error", f"Pandoc failed to convert:\n{e}")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error during processing:\n{e}")
            return

        if not mcq_data:
            messagebox.showinfo("No MCQs", "No MCQs found in the document.")
            return

        # Create Excel file & write data
        try:
            self.write_to_excel(mcq_data, excel_file)
            messagebox.showinfo("Success", f"MCQs saved to Excel file: {excel_file}")
            
            # Open the Excel file
            try:
                os.startfile(excel_file)
            except AttributeError:
                # For non-Windows systems
                import platform
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', excel_file))
                else:  # Linux and other Unix-like
                    subprocess.call(('xdg-open', excel_file))
            except Exception:
                # If opening fails, just inform the user where the file is
                pass
                
        except Exception as e:
            messagebox.showerror("Excel Error", f"Failed to create Excel file:\n{e}")
            return

    def extract_images_from_docx(self, docx_file, output_dir):
        """Extract images from the DOCX file (which is a ZIP archive)"""
        try:
            with zipfile.ZipFile(docx_file, 'r') as zip_ref:
                # Find all image files in the archive
                for file_info in zip_ref.infolist():
                    if file_info.filename.startswith('word/media/'):
                        # Extract the image file
                        image_data = zip_ref.read(file_info.filename)
                        # Get just the filename part
                        image_filename = os.path.basename(file_info.filename)
                        # Save to the output directory
                        with open(os.path.join(output_dir, image_filename), 'wb') as img_file:
                            img_file.write(image_data)
        except Exception as e:
            messagebox.showwarning("Image Extraction", f"Error extracting images: {e}")

    def image_to_base64(self, image_path):
        """Convert an image file to base64 string"""
        try:
            if not os.path.exists(image_path):
                return ""
            
            # Open the image and resize if needed
            with Image.open(image_path) as img:
                # Optionally resize large images
                max_size = (800, 600)  # Maximum dimensions
                if img.width > max_size[0] or img.height > max_size[1]:
                    img.thumbnail(max_size, Image.LANCZOS)
                
                # Convert to base64
                buffer = io.BytesIO()
                # Determine the format based on the file extension
                format_ext = os.path.splitext(image_path)[1].lower().lstrip('.')
                if format_ext not in ['jpg', 'jpeg', 'png', 'gif']:
                    format_ext = 'png'  # Default to PNG for unknown formats
                
                # Save to buffer
                img.save(buffer, format=format_ext.upper())
                img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
                
                # Create data URL
                return f"data:image/{format_ext};base64,{img_str}"
        except Exception as e:
            print(f"Error converting image to base64: {e}")
            return ""

    # ---------------------------------------------------------------------
    # Helper Methods
    # ---------------------------------------------------------------------

    def parse_bracket_tokens(self, text):
        """
        Find any occurrences of '{[}...{]}' in the line, e.g.:
        'সকল মূলদ ...? {[}য. বো. ২০১৫{]} {[}টপিক: মূলদ ও অমূলদ সংখ্যা{]}'
        and extract them as board_institute or topic.
        
        Returns:
            (base_text, board_institute, topic)
        """
        board_institute = ""
        topic = ""

        # Regex to find bracket blocks like: {[}some text{]}
        bracket_pattern = re.compile(r'\{\[}(.*?)\{]}')
        matches = bracket_pattern.findall(text)  # list of contents inside {[}...{]}

        # Remove them from the original line so they're not in the question text
        base_text = bracket_pattern.sub('', text).strip()

        # Now parse each bracket's content to see if it starts with "টপিক:"
        for m in matches:
            stripped_m = m.strip()
            if stripped_m.startswith("টপিক:"):
                # example: "টপিক: মূলদ ও অমূলদ সংখ্যা"
                topic = stripped_m.replace("টপিক:", "").strip()
            else:
                # otherwise, assume it's Board/Institution
                board_institute = stripped_m

        return base_text, board_institute, topic

    def parse_latex_for_mcqs(self, latex_file, images_dir):
        """
        Reads the LaTeX line by line, searching for the structure:

            {Serial}. {Question text} (may contain bracket tokens in the same line)
            {possibly bracket tokens on subsequent lines}
            ক. {Option A}
            খ. {Option B}
            গ. {Option C}
            ঘ. {Option D}
            উত্তরঃ [Answer]
            ব্যাখ্যাঃ [Explanation]

        Returns a list of rows for the final spreadsheet.
        """

        with open(latex_file, "r", encoding="utf-8") as f:
            lines = f.readlines()

        # Regex for "Serial. question"
        re_question = re.compile(r'^([০-৯]+)[.,)।]\s*(.*)', re.UNICODE)
        # Regex for options: "ক. ..."
        re_option = re.compile(r'^([ক-ঘ])[.)]\s+(.*)$', re.UNICODE)
        # Regex for answer: "উত্তরঃ ..."
        re_answer = re.compile(r'^উত্তর[:ঃ]\s+(.*)$', re.UNICODE)
        # Regex for explanation: "ব্যাখ্যাঃ ..."
        re_explanation = re.compile(r'^ব্যাখ্যাঃ\s+(.*)$', re.UNICODE)
        # Regex for image inclusion
        re_image = re.compile(r'\\includegraphics(?:\[.*?\])?\{(.*?)\}')

        mcq_data = []

        # Temporary state
        serial_number = ""
        question_text = ""
        question_img = ""
        board_institute = ""
        topic = ""
        options = {"ক": "", "খ": "", "গ": "", "ঘ": ""}
        options_img = {"ক": "", "খ": "", "গ": "", "ঘ": ""}
        answer = ""
        explanation = ""
        explanation_img = ""
        hint = ""
        hint_img = ""
        current_section = None  # Track which section we're in

        def commit_mcq():
            """Append the current MCQ to mcq_data, then reset."""
            if serial_number:
                # Map Bengali option letters to English
                option_map = {"ক": "A", "খ": "B", "গ": "C", "ঘ": "D"}
                answer_eng = ""
                for bn_letter in answer:
                    if bn_letter in option_map:
                        answer_eng += option_map[bn_letter]
                
                mcq_data.append([
                    serial_number.strip(),      # Serial as a temp ID
                    question_text.strip(),      # Question
                    question_img,               # Question image base64
                    topic.strip(),              # Topic
                    board_institute.strip(),    # Board/Institute
                    options["ক"].strip(),       # Option A
                    options_img["ক"],           # Option A image base64
                    options["খ"].strip(),       # Option B
                    options_img["খ"],           # Option B image base64
                    options["গ"].strip(),       # Option C
                    options_img["গ"],           # Option C image base64
                    options["ঘ"].strip(),       # Option D
                    options_img["ঘ"],           # Option D image base64
                    answer_eng.strip() if answer_eng else answer.strip(), # Answer
                    explanation.strip(),        # Explanation
                    explanation_img,            # Explanation image base64
                    hint.strip(),               # Hint
                    hint_img                    # Hint image base64
                ])

        for line in lines:
            text = line.strip()
            if not text:
                continue

            # Process equations based on user preference
            if self.preserve_equations.get():
                # Keep the $ symbols but clean up LaTeX commands that may cause issues
                text = self.clean_latex_commands(text)
            else:
                # Convert LaTeX equations to Unicode
                text = self.convert_inline_equations_to_unicode(text)

            # Check for images in the line
            img_matches = re_image.findall(text)
            img_base64 = ""
            
            # If images found, convert to base64
            if img_matches:
                for img_path in img_matches:
                    # Extract just the filename part (might be media/image1.png)
                    img_filename = os.path.basename(img_path)
                    # Construct full path to the extracted image
                    full_img_path = os.path.join(images_dir, img_filename)
                    
                    if os.path.exists(full_img_path):
                        img_base64 = self.image_to_base64(full_img_path)
                
                # Remove the image reference from the text
                text = re_image.sub('', text).strip()
                
            # Check if line is "Serial. question"
            mq = re_question.match(text)
            if mq:
                # If we already have an MCQ in progress, commit it before starting a new one
                if serial_number:
                    commit_mcq()

                # Start a new MCQ
                serial_number = mq.group(1)
                q_text = mq.group(2)
                current_section = "question"  # Set current section

                # Extract bracket tokens from the question line
                base_text, new_board, new_topic = self.parse_bracket_tokens(q_text)
                question_text = base_text

                # If bracket tokens found, store them
                board_institute = new_board
                topic = new_topic

                # Store image if found
                if img_base64:
                    question_img = img_base64

                # Reset for new MCQ
                options = {"ক": "", "খ": "", "গ": "", "ঘ": ""}
                options_img = {"ক": "", "খ": "", "গ": "", "ঘ": ""}
                answer = ""
                explanation = ""
                explanation_img = ""
                hint = ""
                hint_img = ""
                continue

            # If we have an active MCQ, we might find bracket tokens or options/answer
            if serial_number:
                # Extract bracket tokens from this line
                base_text, new_board, new_topic = self.parse_bracket_tokens(text)

                # If new board_institute or topic found, store them
                if new_board:
                    board_institute = new_board
                if new_topic:
                    topic = new_topic

                # Check if this line is an option
                mo = re_option.match(base_text)
                if mo:
                    letter = mo.group(1)
                    opt_text = mo.group(2)
                    current_section = f"option_{letter}"
                    options[letter] = opt_text
                    # Store image if found
                    if img_base64:
                        options_img[letter] = img_base64
                    continue

                # Check if it's an answer line
                ma = re_answer.match(base_text)
                if ma:
                    answer_text = ma.group(1)
                    current_section = "answer"
                    answer = answer_text
                    continue

                # Check if it's an explanation line
                me = re_explanation.match(base_text)
                if me:
                    explanation_text = me.group(1)
                    current_section = "explanation"
                    explanation = explanation_text
                    # Store image if found
                    if img_base64:
                        explanation_img = img_base64
                    continue

                # Otherwise, it's additional text for the current section
                if current_section == "question":
                    question_text += " " + base_text
                    # If image found
                    if img_base64 and not question_img:
                        question_img = img_base64
                elif current_section == "explanation":
                    explanation += " " + base_text
                    # If image found
                    if img_base64 and not explanation_img:
                        explanation_img = img_base64
                elif current_section and current_section.startswith("option_"):
                    letter = current_section[-1]
                    options[letter] += " " + base_text
                    # If image found
                    if img_base64 and not options_img[letter]:
                        options_img[letter] = img_base64
                elif current_section == "hint":
                    hint += " " + base_text
                    # If image found
                    if img_base64 and not hint_img:
                        hint_img = img_base64

        # End of file => commit last MCQ if needed
        if serial_number:
            commit_mcq()

        return mcq_data

    def clean_latex_commands(self, text):
        """
        Cleans up potentially problematic LaTeX commands but preserves $ signs
        and the math content.
        """
        # This will preserve the $ delimiters and the content between them
        # but clean up some LaTeX commands that might cause parsing issues
        
        # Handle common LaTeX environments that might be in the text
        text = re.sub(r'\\begin\{.*?\}', '', text)
        text = re.sub(r'\\end\{.*?\}', '', text)
        
        # Clean up some excessive whitespace in equations
        text = re.sub(r'\$\s+', '$', text)
        text = re.sub(r'\s+\$', '$', text)
        
        return text

    def convert_inline_equations_to_unicode(self, text):
        """
        Finds $...$ or $$...$$ blocks in 'text' and converts the inside to naive Unicode.
        """
        pattern = re.compile(r'(\${1,2})(.*?)(\1)', re.DOTALL)

        def replacer(m):
            eq_content = m.group(2).strip()
            return self.convert_latex_to_unicode(eq_content)

        return pattern.sub(replacer, text)

    def convert_latex_to_unicode(self, eq_text):
        """
        Naive approach: \alpha->α, x^2->x², x_2->x₂, etc.
        """
        # Common Greek letters
        greek_map = {
            r'\\alpha': 'α',
            r'\\beta': 'β',
            r'\\gamma': 'γ',
            r'\\delta': 'δ',
            r'\\theta': 'θ',
            r'\\mu': 'μ',
            r'\\pi': 'π',
            r'\\sigma': 'σ',
            r'\\phi': 'φ',
            r'\\omega': 'ω'
        }
        for latex_g, uni_g in greek_map.items():
            eq_text = eq_text.replace(latex_g, uni_g)

        # x^2 -> x²
        eq_text = re.sub(
            r'([A-Za-z0-9])\^([A-Za-z0-9])',
            lambda m: m.group(1) + self.to_superscript(m.group(2)),
            eq_text
        )
        # x_2 -> x₂
        eq_text = re.sub(
            r'([A-Za-z0-9])_([A-Za-z0-9])',
            lambda m: m.group(1) + self.to_subscript(m.group(2)),
            eq_text
        )

        # Common math symbols
        eq_text = eq_text.replace(r'\times', '×')
        eq_text = eq_text.replace(r'\cdot', '·')
        eq_text = eq_text.replace(r'\pm', '±')
        eq_text = eq_text.replace(r'\approx', '≈')
        eq_text = eq_text.replace(r'\neq', '≠')
        # optional spacing for '='
        eq_text = eq_text.replace(r'=', ' = ')

        return eq_text.strip()

    def to_superscript(self, char):
        supers = {
            '0': '⁰', '1': '¹', '2': '²', '3': '³',
            '4': '⁴', '5': '⁵', '6': '⁶', '7': '⁷',
            '8': '⁸', '9': '⁹',
            'n': 'ⁿ', 'i': 'ⁱ', '+': '⁺', '-': '⁻'
        }
        return supers.get(char, '^' + char)

    def to_subscript(self, char):
        subs = {
            '0': '₀', '1': '₁', '2': '₂', '3': '₃',
            '4': '₄', '5': '₅', '6': '₆', '7': '₇',
            '8': '₈', '9': '₉',
            '+': '₊', '-': '₋', '=': '₌', '(': '₍', ')': '₎'
        }
        return subs.get(char, '_' + char)

    def write_to_excel(self, mcq_data, excel_file):
        """Create a new Excel file and write MCQs data to it with the specified columns."""
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MCQs"

        # Define header based on the requested columns
        header = [
            "QuestionID",
            "Serial",
            "Class",
            "Subject",
            "Chapter",
            "Topic",
            "Question",
            "Ques_img",
            "OptionA",
            "OptionA_IMG",
            "OptionB",
            "OptionB_IMG",
            "OptionC",
            "OptionC_IMG",
            "OptionD",
            "OptionD_IMG",
            "Answer",
            "Explaination",
            "Explaination_IMG",
            "Hint",
            "Hint_img",
            "Difficulty_level",
            "Reference_Board/Institute",
            "Reference"
        ]

        # Write header row
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Write data rows - map the data to the new structure
        for row_num, row_data in enumerate(mcq_data, 2):
            # Data is now in format:
            # [serial, question, question_img, topic, board_inst, 
            #  option_a, option_a_img, option_b, option_b_img, option_c, option_c_img, option_d, option_d_img,
            #  answer, explanation, explanation_img, hint, hint_img]
            
            # Extract data
            serial_number = row_data[0]
            question_text = row_data[1]
            question_img = row_data[2]
            topic = row_data[3]
            board_institute = row_data[4]
            option_a = row_data[5]
            option_a_img = row_data[6]
            option_b = row_data[7]
            option_b_img = row_data[8]
            option_c = row_data[9]
            option_c_img = row_data[10]
            option_d = row_data[11]
            option_d_img = row_data[12]
            answer = row_data[13]
            explanation = row_data[14]
            explanation_img = row_data[15]
            hint = row_data[16]
            hint_img = row_data[17]
            
            # Create row for Excel with all columns
            new_row_data = [
                f"Q{serial_number}",      # QuestionID - derive from serial
                serial_number,            # Serial
                "",                       # Class - leave empty
                "",                       # Subject - leave empty
                "",                       # Chapter - leave empty
                topic,                    # Topic
                question_text,            # Question
                question_img,             # Ques_img
                option_a,                 # OptionA
                option_a_img,             # OptionA_IMG
                option_b,                 # OptionB
                option_b_img,             # OptionB_IMG
                option_c,                 # OptionC
                option_c_img,             # OptionC_IMG
                option_d,                 # OptionD
                option_d_img,             # OptionD_IMG
                answer,                   # Answer
                explanation,              # Explaination
                explanation_img,          # Explaination_IMG
                hint,                     # Hint
                hint_img,                 # Hint_img
                "",                       # Difficulty_level - leave empty
                board_institute,          # Reference_Board/Institute
                ""                        # Reference - leave empty
            ]
            
            # Write the row to the worksheet
            for col_num, cell_value in enumerate(new_row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                # Apply wrapping to text columns
                if col_num in [7, 9, 11, 13, 15, 18, 20]:  # Text columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-adjust column widths based on content (except for image columns)
        for col_idx, col in enumerate(ws.columns, 1):
            # Skip image columns for width adjustment
            if col_idx in [8, 10, 12, 14, 16, 19, 21]:
                continue
                
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column].width = adjusted_width
            
        # Set fixed width for image columns
        image_columns = [8, 10, 12, 14, 16, 19, 21]
        for col_idx in image_columns:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 30

        # Save the workbook
        wb.save(excel_file)


def main():
    root = tk.Tk()
    app = DocxToExcelPandocGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()