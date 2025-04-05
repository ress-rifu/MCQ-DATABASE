import os
import re
import subprocess
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment
import base64
import zipfile
from PIL import Image
import io
import shutil

class DocxToExcelPandocGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("Docx to Excel (with Pandoc)")

        # StringVars for user inputs
        self.docx_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.preserve_equations = tk.BooleanVar(value=True)

        # New StringVars for metadata
        self.class_name = tk.StringVar()
        self.subject_name = tk.StringVar()
        self.chapter_name = tk.StringVar()

        # Layout
        self.create_widgets()

    def create_widgets(self):
        # Row 0: Word doc
        tk.Label(self.master, text="Word Document:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(self.master, textvariable=self.docx_path, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(self.master, text="Browse", command=self.browse_docx).grid(row=0, column=2, padx=5, pady=5)

        # Row 1: Excel output
        tk.Label(self.master, text="Excel Output:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        tk.Entry(self.master, textvariable=self.excel_path, width=50).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(self.master, text="Browse", command=self.browse_excel).grid(row=1, column=2, padx=5, pady=5)

        # Row 2: Metadata fields - Class
        class_label = tk.Label(self.master, text="Class:")
        class_label.grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.add_tooltip(class_label, "Enter the class name (e.g., '10', 'HSC'). Will be applied to all questions.")
        class_entry = tk.Entry(self.master, textvariable=self.class_name, width=50)
        class_entry.grid(row=2, column=1, padx=5, pady=5)
        self.add_tooltip(class_entry, "Enter the class name (e.g., '10', 'HSC'). Will be applied to all questions.")
        
        # Row 3: Metadata fields - Subject
        subject_label = tk.Label(self.master, text="Subject:")
        subject_label.grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.add_tooltip(subject_label, "Enter the subject name (e.g., 'Physics', 'Mathematics'). Will be applied to all questions.")
        subject_entry = tk.Entry(self.master, textvariable=self.subject_name, width=50)
        subject_entry.grid(row=3, column=1, padx=5, pady=5)
        self.add_tooltip(subject_entry, "Enter the subject name (e.g., 'Physics', 'Mathematics'). Will be applied to all questions.")
        
        # Row 4: Metadata fields - Chapter
        chapter_label = tk.Label(self.master, text="Chapter:")
        chapter_label.grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.add_tooltip(chapter_label, "Enter the chapter name (e.g., 'Kinematics', 'Algebra'). Will be applied to all questions.")
        chapter_entry = tk.Entry(self.master, textvariable=self.chapter_name, width=50)
        chapter_entry.grid(row=4, column=1, padx=5, pady=5)
        self.add_tooltip(chapter_entry, "Enter the chapter name (e.g., 'Kinematics', 'Algebra'). Will be applied to all questions.")

        # Row 5: Equation options
        equations_cb = tk.Checkbutton(self.master, text="Keep equations in pure linear LaTeX format (like MS Word)", 
                      variable=self.preserve_equations)
        equations_cb.grid(row=5, column=1, padx=5, pady=5, sticky="w")
        self.add_tooltip(equations_cb, "When checked, keeps equations in LaTeX format ($...$). Uncheck to attempt conversion to Unicode.")

        # Row 6: Convert button
        convert_btn = tk.Button(self.master, text="Convert & Save", command=self.on_convert_click, width=20)
        convert_btn.grid(row=6, column=1, pady=15)
        self.add_tooltip(convert_btn, "Convert the Word document to Excel with MCQs")

    def browse_docx(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Word Documents", "*.docx"), ("All Files", "*.*")]
        )
        if file_path:
            self.docx_path.set(file_path)
            # Auto-generate Excel output path
            if not self.excel_path.get():
                base_path = os.path.splitext(file_path)[0]
                self.excel_path.set(f"{base_path}.xlsx")

    def browse_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.excel_path.set(file_path)

    def on_convert_click(self):
        docx_file = self.docx_path.get().strip()
        excel_file = self.excel_path.get().strip()

        # Basic checks
        if not docx_file or not os.path.exists(docx_file):
            messagebox.showerror("Error", "Please select a valid .docx file.")
            return
        if not excel_file:
            messagebox.showerror("Error", "Please specify the Excel output file.")
            return

        # Extract images from the docx file
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                print(f"Created temporary directory: {tmpdir}")
                # Extract DOCX contents (it's a ZIP file)
                images_dir = os.path.join(tmpdir, "media")
                os.makedirs(images_dir, exist_ok=True)
                
                # Extract images from DOCX
                self.extract_images_from_docx(docx_file, images_dir)
                print(f"Extracted images to {images_dir}")
                
                # Convert docx -> .tex using pandoc
                tex_path = os.path.join(tmpdir, "converted.tex")
                cmd = ["pandoc", docx_file, "-o", tex_path]
                print(f"Running pandoc command: {' '.join(cmd)}")
                result = subprocess.run(cmd, check=True, capture_output=True, text=True)
                print(f"Pandoc conversion completed. Output file: {tex_path}")
                
                if not os.path.exists(tex_path) or os.path.getsize(tex_path) == 0:
                    print("Error: Generated .tex file is empty or doesn't exist.")
                    messagebox.showerror("Conversion Error", 
                                        "Pandoc failed to generate a proper .tex file. This might be due to formatting issues in your Word document.")
                    return

                # Also create an HTML version for web display
                html_path = os.path.join(tmpdir, "converted.html")
                html_cmd = ["pandoc", docx_file, "-o", html_path]
                print(f"Running pandoc HTML command: {' '.join(html_cmd)}")
                html_result = subprocess.run(html_cmd, check=True, capture_output=True, text=True)
                
                # Extract tables from the HTML for web display
                tables_html = self.extract_tables_from_html(html_path)
                print(f"Extracted {len(tables_html)} tables from the document")
                
                # Save tables to HTML file
                if tables_html:
                    tables_output_path = os.path.splitext(excel_file)[0] + "_tables.html"
                    with open(tables_output_path, "w", encoding="utf-8") as f:
                        f.write("<html><head><title>Extracted Tables</title>")
                        f.write("<style>table {border-collapse: collapse; width: 100%; margin-bottom: 20px;}")
                        f.write("th, td {border: 1px solid #ddd; padding: 8px; text-align: left;}")
                        f.write("th {background-color: #f2f2f2;}</style></head><body>")
                        f.write("<h1>Tables Extracted from Document</h1>")
                        for i, table_html in enumerate(tables_html):
                            f.write(f"<h2>Table {i+1}</h2>")
                            f.write(table_html)
                        f.write("</body></html>")
                    print(f"Saved {len(tables_html)} tables to {tables_output_path}")

                # Parse the generated .tex for MCQs in the desired structure
                mcq_data = self.parse_latex_for_mcqs(tex_path, images_dir)
                
        except FileNotFoundError:
            messagebox.showerror("Pandoc Error", "pandoc not found. Please install pandoc.")
            return
        except subprocess.CalledProcessError as e:
            print(f"Pandoc command failed with error: {e}")
            print(f"Error output: {e.stderr}")
            messagebox.showerror("Pandoc Error", f"Pandoc failed to convert:\n{e}\n\nStderr: {e.stderr}")
            return
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Unexpected error during processing:\n{e}")
            return

        if not mcq_data:
            print("No MCQs were detected in the document.")
            error_msg = (
                "No MCQs found in the document. Please ensure your document follows one of these patterns:\n\n"
                "Pattern 1 (General MCQ):\n"
                "১. Question text    (Serial numbers can be in Bengali ১, ২, ৩... or English 1, 2, 3...)\n"
                "[টপিক: Topic] or [Topic: Your topic] or [Topic information]\n"
                "[Easy] or [Medium] or [Hard] or [Difficulty: Level]\n"
                "[Board-Year] or [Reference: Source] or [Institute: Name]\n"
                "ক. Option A\n"
                "খ. Option B\n"
                "গ. Option C\n"
                "ঘ. Option D\n"
                "উত্তর: Answer\n"
                "[Hint: Hint text]\n"
                "[Explaination: Explanation text]\n\n"
                "Pattern 2 (MCQs with multiple choice answers):\n"
                "১. Question text with statements    (Serial numbers can be in Bengali ১, ২, ৩... or English 1, 2, 3...)\n"
                "[টপিক: Topic] or [Topic: Your topic] or [Topic information]\n"
                "[Easy] or [Medium] or [Hard] or [Difficulty: Level]\n"
                "[Board-Year] or [Reference: Source] or [Institute: Name]\n"
                "i. Statement 1\n"
                "ii. Statement 2\n"
                "iii. Statement 3\n"
                "নিচের কোনটি সঠিক?\n"
                "ক. i and ii\n"
                "খ. i and iii\n"
                "গ. ii and iii\n"
                "ঘ. i, ii and iii\n"
                "উত্তর: Answer\n"
                "[Hint: Hint text]\n"
                "[Explaination: Explanation text]"
            )
            messagebox.showinfo("No MCQs", error_msg)
            return

        # Create Excel file & write data
        try:
            self.write_to_excel(mcq_data, excel_file)
            
            message = f"{len(mcq_data)} MCQs saved to Excel file: {excel_file}"
            if hasattr(self, 'tables_found') and self.tables_found:
                tables_output_path = os.path.splitext(excel_file)[0] + "_tables.html"
                message += f"\n{self.tables_found} tables extracted to: {tables_output_path}"
                
            messagebox.showinfo("Success", message)
            
            # Open the Excel file
            try:
                os.startfile(excel_file)
            except AttributeError:
                # For non-Windows systems
                import platform
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', excel_file))
                else:  # Linux and other Unix-like
                    subprocess.call(('xdg-open', excel_file))
            except Exception:
                # If opening fails, just inform the user where the file is
                pass
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Excel Error", f"Failed to create Excel file:\n{e}")
            return

    def extract_images_from_docx(self, docx_file, output_dir):
        """Extract images from the DOCX file (which is a ZIP archive)"""
        try:
            with zipfile.ZipFile(docx_file, 'r') as zip_ref:
                # Find all image files in the archive
                for file_info in zip_ref.infolist():
                    if file_info.filename.startswith('word/media/'):
                        # Extract the image file
                        image_data = zip_ref.read(file_info.filename)
                        # Get just the filename part
                        image_filename = os.path.basename(file_info.filename)
                        # Save to the output directory
                        with open(os.path.join(output_dir, image_filename), 'wb') as img_file:
                            img_file.write(image_data)
        except Exception as e:
            messagebox.showwarning("Image Extraction", f"Error extracting images: {e}")

    def image_to_base64(self, image_path):
        """Convert an image file to base64 string"""
        try:
            if not os.path.exists(image_path):
                return ""
            
            # Open the image and resize if needed
            with Image.open(image_path) as img:
                # Optionally resize large images
                max_size = (800, 600)  # Maximum dimensions
                if img.width > max_size[0] or img.height > max_size[1]:
                    img.thumbnail(max_size, Image.LANCZOS)
                
                # Convert to base64
                buffer = io.BytesIO()
                # Determine the format based on the file extension
                format_ext = os.path.splitext(image_path)[1].lower().lstrip('.')
                if format_ext not in ['jpg', 'jpeg', 'png', 'gif']:
                    format_ext = 'png'  # Default to PNG for unknown formats
                
                # Save to buffer
                img.save(buffer, format=format_ext.upper())
                img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
                
                # Create data URL
                return f"data:image/{format_ext};base64,{img_str}"
        except Exception as e:
            print(f"Error converting image to base64: {e}")
            return ""

    # ---------------------------------------------------------------------
    # Helper Methods
    # ---------------------------------------------------------------------

    def parse_bracket_tokens(self, text):
        """
        Find any occurrences of '{[}...{]}' in the line, e.g.:
        'সকল মূলদ ...? {[}য. বো. ২০১৫{]} {[}টপিক: মূলদ ও অমূলদ সংখ্যা{]}'
        and extract them as board_institute or topic.
        
        Returns:
            (base_text, board_institute, topic)
        """
        board_institute = ""
        topic = ""

        # First, temporarily protect equation content within $ signs
        protected_text = text
        equations = []
        
        # Find and protect $...$ expressions
        eq_pattern = re.compile(r'(\$+)(.*?)(\1)', re.DOTALL)
        eq_matches = eq_pattern.findall(text)
        
        for i, match in enumerate(eq_matches):
            delim, content, _ = match
            placeholder = f"__EQ_{i}__"
            equations.append(f"{delim}{content}{delim}")
            protected_text = protected_text.replace(f"{delim}{content}{delim}", placeholder)

        # Regex to find bracket blocks like: {[}some text{]}
        bracket_pattern = re.compile(r'\{\[}(.*?)\{]}')
        matches = bracket_pattern.findall(protected_text)  # list of contents inside {[}...{]}

        # Remove them from the original line so they're not in the question text
        base_text = bracket_pattern.sub('', protected_text).strip()

        # Now parse each bracket's content to see if it starts with "টপিক:"
        for m in matches:
            stripped_m = m.strip()
            if stripped_m.startswith("টপিক:"):
                # example: "টপিক: মূলদ ও অমূলদ সংখ্যা"
                topic = stripped_m.replace("টপিক:", "").strip()
            else:
                # otherwise, assume it's Board/Institution
                board_institute = stripped_m

        # Restore equations in base_text
        for i, eq in enumerate(equations):
            base_text = base_text.replace(f"__EQ_{i}__", eq)

        return base_text, board_institute, topic

    def parse_latex_for_mcqs(self, latex_file, images_dir):
        """
        Reads the LaTeX line by line, searching for the structure of both pattern types:

        Pattern 1 (General MCQ):
            {Serial}. {Question text} 
            [টপিক: {Topic}]
            [Difficulty]
            [Board/Institute info]
            ক. {Option A}
            খ. {Option B}
            গ. {Option C}
            ঘ. {Option D}
            উত্তরঃ [Answer]
            [Hint: {Hint}]
            [Explaination: {Explanation}]

        Pattern 2 (MCQs with multiple choice answers):
            {Serial}. {Question text with statements i, ii, iii} 
            [টপিক: {Topic}]
            [Difficulty]
            [Board/Institute info]
            নিচের কোনটি সঠিক?
            ক. {Option A}
            খ. {Option B}
            গ. {Option C}
            ঘ. {Option D}
            উত্তরঃ [Answer]
            [Hint: {Hint}]
            [Explaination: {Explanation}]

        Returns a list of rows for the final spreadsheet.
        """

        with open(latex_file, "r", encoding="utf-8") as f:
            lines = f.readlines()

        # Print first few lines for debugging
        print("First few lines of the latex file:")
        for i, line in enumerate(lines[:10]):
            print(f"Line {i+1}: {line.strip()}")

        # Regex for "Serial. question" - Bengali numerals
        re_question_bn = re.compile(r'^([০-৯]+)[.,)।]\s*(.*)', re.UNICODE)
        # Regex for "Serial. question" - English numerals
        re_question_en = re.compile(r'^(\d+)[.,)।]\s*(.*)', re.UNICODE)
        # Regex for options: "ক. ..."
        re_option = re.compile(r'^([ক-ঘ])[.)]\s+(.*)$', re.UNICODE)
        # Regex for answer: "উত্তরঃ ..."
        re_answer = re.compile(r'^উত্তর[:ঃ]\s+(.*)$', re.UNICODE)
        # Regex for explanation: "[Explaination: ...]"
        re_explanation = re.compile(r'\[Explaination:\s+(.*?)(?:\]|$)', re.UNICODE)
        # Regex for hint: "[Hint: ...]"
        re_hint = re.compile(r'\[Hint:\s+(.*?)(?:\]|$)', re.UNICODE)
        
        # Updated regex patterns for topic, difficulty, and board/institute
        # Regex for difficulty: "[Easy]", "[Medium]", "[Hard]", or any text in brackets
        re_difficulty = re.compile(r'\[(Easy|Medium|Hard|.*?Difficulty.*?)\]', re.UNICODE | re.IGNORECASE)
        # Regex for board/institute: "[Board-Year]" or any text containing "Board" in brackets
        re_board = re.compile(r'\[(.*?(?:Board|Institute|Reference).*?)\]', re.UNICODE | re.IGNORECASE)
        # Regex for topic: "[টপিক: ...]" or any text starting with "Topic:" in brackets
        re_topic = re.compile(r'\[(?:টপিক[:ঃ]|Topic:)\s+(.*?)(?:\]|$)', re.UNICODE | re.IGNORECASE)
        # Alternative topic pattern (just in brackets)
        re_topic_alt = re.compile(r'\[(.*?(?:Topic|Subject|Chapter).*?)\]', re.UNICODE | re.IGNORECASE)
        
        # Updated specific pattern for Bengali topics - with non-greedy matching to get full content
        re_topic_bengali = re.compile(r'\[টপিক[:ঃ]\s+([\s\S]*?)\]', re.UNICODE)
        
        # Regex for image inclusion
        re_image = re.compile(r'\\includegraphics(?:\[.*?\])?\{(.*?)\}')
        # Regex for multiple choice statements (i, ii, iii)
        re_statement = re.compile(r'(?:^|\n)\s*(i{1,3})\.\s*(.*?)(?=$|\n\s*i{1,3}\.|\n\s*নিচের)', re.UNICODE | re.DOTALL)
        # Regex for "নিচের কোনটি সঠিক?" to detect Pattern 2
        re_which_correct = re.compile(r'নিচের\s+কোনটি\s+সঠিক\s*\?', re.UNICODE)

        mcq_data = []

        # First pass: collect all lines into a single text
        full_text = " ".join([line.strip() for line in lines if line.strip()])
        
        # Print first 100 characters for debugging
        print(f"Full text first 100 chars: {full_text[:100]}")
        
        # Search for Bengali numerals in the text to confirm they exist
        bengali_digits_pattern = re.compile(r'[০-৯]+')
        bengali_matches = bengali_digits_pattern.findall(full_text)
        if bengali_matches:
            print(f"Found Bengali digits in text: {', '.join(bengali_matches[:10])}")
        else:
            print("No Bengali digits found in the text")
        
        # More robust pattern for Bengali serial numbers with various separators
        bengali_serial_pattern = r'(?:^|\s)([০-৯]+)[\.|\,|\)|।|:|\\]?[\s]*'
        
        # Try pattern matching with Bengali digits first (priority)
        mcq_blocks = re.split(bengali_serial_pattern, full_text)
        
        # If no MCQs found or very few, try with English digits
        if len(mcq_blocks) <= 3:  # Only the original text or just one MCQ found
            print("Few or no Bengali digit MCQs found, trying with English digits...")
            mcq_blocks = re.split(r'(?:^|\s)(\d+)[\.|\,|\)|।|:|\\]?[\s]*', full_text)
        
        # If still no MCQs found, try with even more flexible pattern
        if len(mcq_blocks) <= 3:
            print("Still few or no MCQs found, trying with an even more flexible pattern...")
            # This pattern allows for more variations in separators and spacing
            combined_pattern = r'(?:^|\s)([০-৯\d]+)[\.|\,|\)|।|:|\\]?[\s]*'
            mcq_blocks = re.split(combined_pattern, full_text)
        
        print(f"Found {(len(mcq_blocks)-1)//2} potential MCQ blocks")
        
        # Debug first block if available
        if len(mcq_blocks) > 2:
            print(f"First MCQ serial: {mcq_blocks[1]}")
            print(f"First MCQ content (first 100 chars): {mcq_blocks[2][:100]}...")
        
        # Check if any of the serial numbers are Bengali digits
        bengali_serials = 0
        english_serials = 0
        for i in range(1, len(mcq_blocks), 2):
            if i < len(mcq_blocks):
                if re.match(r'[০-৯]+', mcq_blocks[i]):
                    bengali_serials += 1
                elif re.match(r'\d+', mcq_blocks[i]):
                    english_serials += 1
        
        print(f"Detected {bengali_serials} Bengali serial numbers and {english_serials} English serial numbers")
        
        # Process each MCQ block
        i = 1
        while i < len(mcq_blocks):
            # Get serial number and question text
            serial_number = mcq_blocks[i]
            question_text = mcq_blocks[i+1] if i+1 < len(mcq_blocks) else ""
            
            # Skip if no question text
            if not question_text.strip():
                i += 2
                continue
                
            print(f"Processing MCQ with serial: {serial_number}")
            
            # Check if it's a Bengali serial number and display equivalent English number
            if re.match(r'[০-৯]+', serial_number):
                # Convert Bengali digits to English
                english_serial = ''
                bengali_to_english = {'০':'0', '১':'1', '২':'2', '৩':'3', '৪':'4', 
                                     '৫':'5', '৬':'6', '৭':'7', '৮':'8', '৯':'9'}
                for digit in serial_number:
                    english_serial += bengali_to_english.get(digit, digit)
                print(f"  Bengali serial {serial_number} = English serial {english_serial}")
            
            # Reset variables for new MCQ
            board_institute = ""
            topic = ""
            difficulty = ""
            question_img = ""
            options = {"ক": "", "খ": "", "গ": "", "ঘ": ""}
            options_img = {"ক": "", "খ": "", "গ": "", "ঘ": ""}
            answer = ""
            explanation = ""
            explanation_img = ""
            hint = ""
            hint_img = ""
            is_pattern2 = False
            
            # Check if it's Pattern 2 (has multiple choice statements)
            if re_which_correct.search(question_text):
                is_pattern2 = True
                print(f"MCQ {serial_number} is Pattern 2 (multiple choice)")
            
            # Extract images from question text
            img_matches = re_image.findall(question_text)
            if img_matches:
                for img_path in img_matches:
                    # Extract just the filename part
                    img_filename = os.path.basename(img_path)
                    # Construct full path to the extracted image
                    full_img_path = os.path.join(images_dir, img_filename)
                    
                    if os.path.exists(full_img_path):
                        question_img = self.image_to_base64(full_img_path)
                        # Once we have a question image, we can break
                        break
                
                # Remove image references from text
                question_text = re_image.sub('', question_text)
            
            # Extract topic from the question text - try both patterns
            # First try the Bengali-specific pattern
            topic_match = re_topic_bengali.search(question_text)
            if topic_match:
                topic = topic_match.group(1).strip()
                print(f"Found topic (Bengali pattern): {topic}")
            else:
                # Try general topic pattern
                topic_match = re_topic.search(question_text)
                if topic_match:
                    topic = topic_match.group(1).strip()
                    print(f"Found topic (primary pattern): {topic}")
                else:
                    # Try alternative topic pattern
                    topic_alt_match = re_topic_alt.search(question_text)
                    if topic_alt_match:
                        topic = topic_alt_match.group(1).strip()
                        print(f"Found topic (alternative pattern): {topic}")
            
            # Extract difficulty from the question text
            difficulty_match = re_difficulty.search(question_text)
            if difficulty_match:
                difficulty = difficulty_match.group(1).strip()
                print(f"Found difficulty: {difficulty}")
            
            # Extract board/institute from the question text
            board_match = re_board.search(question_text)
            if board_match:
                board_institute = board_match.group(1).strip()
                print(f"Found board/institute: {board_institute}")
            
            # Extract hint from the question text
            hint_match = re_hint.search(question_text)
            if hint_match:
                hint = hint_match.group(1).strip()
                print(f"Found hint of length: {len(hint)}")
                
                # Look for images in the hint text
                hint_img_matches = re_image.findall(hint)
                if hint_img_matches:
                    for img_path in hint_img_matches:
                        img_filename = os.path.basename(img_path)
                        full_img_path = os.path.join(images_dir, img_filename)
                        
                        if os.path.exists(full_img_path):
                            hint_img = self.image_to_base64(full_img_path)
                            break
                    
                    # Remove image references from hint
                    hint = re_image.sub('', hint)
            
            # Extract explanation from the question text
            explanation_match = re_explanation.search(question_text)
            if explanation_match:
                explanation = explanation_match.group(1).strip()
                print(f"Found explanation of length: {len(explanation)}")
                
                # Look for images in the explanation text
                exp_img_matches = re_image.findall(explanation)
                if exp_img_matches:
                    for img_path in exp_img_matches:
                        img_filename = os.path.basename(img_path)
                        full_img_path = os.path.join(images_dir, img_filename)
                        
                        if os.path.exists(full_img_path):
                            explanation_img = self.image_to_base64(full_img_path)
                            break
                    
                    # Remove image references from explanation
                    explanation = re_image.sub('', explanation)
            
            # Try to find options with two different patterns
            option_found = False
            for option_letter in ["ক", "খ", "গ", "ঘ"]:
                # Pattern 1: Look for options like "ক. Option text"
                option_pattern1 = re.compile(rf"{option_letter}\.\s+(.*?)(?=\s+[ক-ঘ]\.|উত্তর[:ঃ]|\[Hint:|\[Explaination:|$)", re.DOTALL)
                option_match = option_pattern1.search(question_text)
                
                # Pattern 2: Look for options like "ক) Option text" or "ক অপশন টেক্সট"
                if not option_match:
                    option_pattern2 = re.compile(rf"{option_letter}[\)।\s]\s*(.*?)(?=\s+[ক-ঘ][\)।\s]|উত্তর[:ঃ]|\[Hint:|\[Explaination:|$)", re.DOTALL)
                    option_match = option_pattern2.search(question_text)
                
                if option_match:
                    option_found = True
                    option_text = option_match.group(1).strip()
                    print(f"Found option {option_letter}: {option_text[:20]}...")
                    
                    # Look for images in the option text
                    opt_img_matches = re_image.findall(option_text)
                    if opt_img_matches:
                        for img_path in opt_img_matches:
                            img_filename = os.path.basename(img_path)
                            full_img_path = os.path.join(images_dir, img_filename)
                            
                            if os.path.exists(full_img_path):
                                options_img[option_letter] = self.image_to_base64(full_img_path)
                                break
                        
                        # Remove image references from option text
                        option_text = re_image.sub('', option_text)
                    
                    options[option_letter] = option_text
            
            # Skip this MCQ if no options found
            if not option_found:
                print(f"Warning: No options found for MCQ {serial_number}, skipping")
                i += 2
                continue
            
            # Extract answer from the question text - try multiple patterns
            answer_match = re.search(r'উত্তর[:ঃ]\s+(.*?)(?=\s+\[|$)', question_text)
            if not answer_match:
                answer_match = re.search(r'[Aa]nswer[:ঃ]\s+(.*?)(?=\s+\[|$)', question_text)
            
            if answer_match:
                answer = answer_match.group(1).strip()
                print(f"Found answer: {answer}")
            else:
                print(f"Warning: No answer found for MCQ {serial_number}")
            
            # Clean up the question text by removing extracted parts
            cleaned_question = question_text
            # Remove topic - both patterns
            cleaned_question = re.sub(r'\[(?:টপিক[:ঃ]|Topic:)\s+.*?\]', '', cleaned_question, flags=re.IGNORECASE)
            cleaned_question = re.sub(r'\[(.*?(?:Topic|Subject|Chapter).*?)\]', '', cleaned_question, flags=re.IGNORECASE)
            # Remove difficulty
            cleaned_question = re.sub(r'\[(Easy|Medium|Hard|.*?Difficulty.*?)\]', '', cleaned_question, flags=re.IGNORECASE)
            # Remove board/institute
            cleaned_question = re.sub(r'\[(.*?(?:Board|Institute|Reference).*?)\]', '', cleaned_question, flags=re.IGNORECASE)
            # Remove options
            for option_letter in ["ক", "খ", "গ", "ঘ"]:
                cleaned_question = re.sub(rf"{option_letter}[.)]\s+.*?(?=\s+[ক-ঘ][.)]|উত্তর[:ঃ]|\[Hint:|\[Explaination:|$)", '', cleaned_question, flags=re.DOTALL)
            # Remove answer
            cleaned_question = re.sub(r'উত্তর[:ঃ]\s+.*?(?=\s+\[|$)', '', cleaned_question)
            cleaned_question = re.sub(r'[Aa]nswer[:ঃ]\s+.*?(?=\s+\[|$)', '', cleaned_question)
            # Remove hint
            cleaned_question = re.sub(r'\[Hint:\s+.*?\]', '', cleaned_question)
            # Remove explanation
            cleaned_question = re.sub(r'\[Explaination:\s+.*?\]', '', cleaned_question)
            # Remove the "নিচের কোনটি সঠিক?" text for pattern 2
            if is_pattern2:
                cleaned_question = re.sub(r'নিচের\s+কোনটি\s+সঠিক\s*\?', '', cleaned_question)
            
            # Clean up any excessive whitespace
            if is_pattern2:
                # For pattern 2, preserve newlines but replace multiple spaces with single space
                cleaned_question = re.sub(r' +', ' ', cleaned_question).strip()
                # Make sure there are no more than 2 consecutive newlines
                cleaned_question = re.sub(r'\n{3,}', '\n\n', cleaned_question)
            else:
                # For pattern 1, replace all whitespace with single space
                cleaned_question = re.sub(r'\s+', ' ', cleaned_question).strip()
            
            # Update question text with the cleaned version
            question_text = cleaned_question
            print(f"Cleaned question (first 50 chars): {question_text[:50]}...")
            
            # Process equations based on user preference
            if self.preserve_equations.get():
                # Keep the $ symbols but clean up LaTeX commands that may cause issues
                question_text = self.clean_latex_commands(question_text)
                topic = self.clean_latex_commands(topic)
                board_institute = self.clean_latex_commands(board_institute)
                hint = self.clean_latex_commands(hint)
                explanation = self.clean_latex_commands(explanation)
                for k in options.keys():
                    options[k] = self.clean_latex_commands(options[k])
            else:
                # Convert LaTeX equations to Unicode
                question_text = self.convert_inline_equations_to_unicode(question_text)
                topic = self.convert_inline_equations_to_unicode(topic)
                board_institute = self.convert_inline_equations_to_unicode(board_institute)
                hint = self.convert_inline_equations_to_unicode(hint)
                explanation = self.convert_inline_equations_to_unicode(explanation)
                for k in options.keys():
                    options[k] = self.convert_inline_equations_to_unicode(options[k])
            
            # Convert Bengali answer to English (ক -> A, খ -> B, etc.)
            option_map = {"ক": "A", "খ": "B", "গ": "C", "ঘ": "D"}
            answer_eng = ""
            for bn_letter in answer:
                if bn_letter in option_map:
                    answer_eng += option_map[bn_letter]
            
            # If we successfully mapped the answer, use it; otherwise keep original
            if answer_eng:
                answer = answer_eng
            
            # Add the MCQ to our data
            mcq_data.append([
                serial_number.strip(),      # Serial as a temp ID
                question_text.strip(),      # Question
                question_img,               # Question image base64
                topic.strip(),              # Topic
                difficulty.strip(),         # Difficulty level
                board_institute.strip(),    # Board/Institute
                options["ক"].strip(),       # Option A
                options_img["ক"],           # Option A image base64
                options["খ"].strip(),       # Option B
                options_img["খ"],           # Option B image base64
                options["গ"].strip(),       # Option C
                options_img["গ"],           # Option C image base64
                options["ঘ"].strip(),       # Option D
                options_img["ঘ"],           # Option D image base64
                answer.strip(),             # Answer
                explanation.strip(),        # Explanation
                explanation_img,            # Explanation image base64
                hint.strip(),               # Hint
                hint_img                    # Hint image base64
            ])

            print(f"Successfully added MCQ {serial_number} to dataset")
            
            # Move to the next block
            i += 2

        return mcq_data

    def clean_latex_commands(self, text):
        """
        Cleans up potentially problematic LaTeX commands but preserves $ signs
        and the math content in pure linear format (like MS Word).
        """
        # First, temporarily protect special LaTeX commands that use $ internally
        text = text.replace(r'$\neq$', '__SPECIAL_NEQ__')
        
        # Replace LaTeX equation delimiters \( and \) with $ signs
        text = re.sub(r'\\(\(|\))', '$', text)
        
        # Fix adjacent $ signs (might happen with $$ in the middle)
        text = re.sub(r'\${2,}', '$', text)
        
        # Handle specific case for system of equations with braces (like in the example image)
        # $\left.\ \begin{matrix}-\frac{1}{2}x+y&=-1\\x-2y&=2\\\end{matrix}\right\}$
        system_pattern = r'\\left\.\s*\\begin\{matrix\}(.*?)\\end\{matrix\}\\right\\}'
        if re.search(system_pattern, text, re.DOTALL):
            # Keep this format as is - it's already in the desired linear format with braces
            pass
        else:
            # Handle other LaTeX for system of equations in linear format
            # Convert any matrix or array environment to linear equations
            text = re.sub(r'\\left\\{.*?\\begin\{(matrix|array).*?\}(.*?)\\end\{\1.*?\}', 
                         lambda m: self.linearize_equation_system(m.group(2)), text, flags=re.DOTALL)
        
        # Handle common LaTeX environments that might be in the text but preserve their content
        text = re.sub(r'\\begin\{(equation|align|gather|eqnarray)\*?\}(.*?)\\end\{\1\*?\}', 
                     lambda m: f'${m.group(2).strip()}$', text, flags=re.DOTALL)
        
        # Clean up some excessive whitespace in equations
        text = re.sub(r'\$\s+', '$', text)
        text = re.sub(r'\s+\$', '$', text)
        
        # Handle the specific case from the example
        # For example: $\frac{a_{1}}{a_{2}} = \frac{b_{1}}{b_{2}} $\neq$ \frac{c_{1}}{c_{2}}$
        text = re.sub(r'\$([^$]*?) \$\\neq\$ ([^$]*?)\$', r'$\1 \\neq \2$', text)
        
        # Remove \textbf and similar LaTeX markup tags
        text = re.sub(r'\\textbf\{([^}]*?)\}', r'\1', text)
        text = re.sub(r'\\textit\{([^}]*?)\}', r'\1', text)
        text = re.sub(r'\\emph\{([^}]*?)\}', r'\1', text)
        
        # Remove extra brackets around normal text
        text = re.sub(r'\{\[}([^{}\[\]]*?)\{\]}', r'\1', text)
        text = re.sub(r'\textbf\{\{([^{}]*?)\}\}', r'\1', text)
        
        # Handle the specific bracket pattern from the example
        # \textbf{{[}}টপিক\textbf{:} ... \textbf{{]}}
        text = re.sub(r'\\textbf\{\{\[}}(.*?)\\textbf\{\{\]}}', r'[\1]', text)
        text = re.sub(r'\\textbf\{:\}', r':', text)
        text = re.sub(r'\\textbf\{\?}', r'?', text)
        
        # Remove extra brackets that might interfere with equation display
        text = re.sub(r'\{\\\[}(.*?)\{\\\]}', r'\1', text)
        text = re.sub(r'\{\\(.*?)\\}', r'\1', text)
        
        # Restore protected special commands
        text = text.replace('__SPECIAL_NEQ__', r'\neq')
        
        # Clean up slashes before and after equation delimiters
        # Remove unnecessary backslashes and forward slashes before $
        text = re.sub(r'[\/\\]+\s*\$', ' $', text)
        # Remove unnecessary backslashes and forward slashes after $
        text = re.sub(r'\$\s*[\/\\]+', '$ ', text)
        
        # Add spaces before and after every equation if not already present
        # First, find all equation patterns (text between $ signs)
        equation_pattern = re.compile(r'(\$[^$]+?\$)')
        
        # Function to add spaces only if needed
        def add_spaces_to_eq(match):
            eq = match.group(1)
            # Check for spaces before and after
            prefix = ' ' if not eq.startswith(' $') else ''
            suffix = ' ' if not eq.endswith('$ ') else ''
            return f"{prefix}{eq}{suffix}"
            
        # Apply the spacing function
        text = equation_pattern.sub(add_spaces_to_eq, text)
        
        # Fix any cases where we might have added too many spaces
        text = re.sub(r'\s{2,}', ' ', text)
        
        return text.strip()
        
    def linearize_equation_system(self, matrix_content):
        """
        Convert a LaTeX matrix/array representation of a system of equations
        to a linear format like MS Word would display.
        
        Example:
        Input: -\frac{1}{2}x+y&=-1\\x-2y&=2
        Output: -\frac{1}{2}x+y=-1, x-2y=2
        """
        # Replace newline markers with commas
        linear = re.sub(r'\\\\', ', ', matrix_content)
        
        # Replace alignment markers with appropriate symbols
        linear = re.sub(r'&=', '=', linear)
        linear = re.sub(r'&<', '<', linear)
        linear = re.sub(r'&>', '>', linear)
        
        return linear

    def convert_inline_equations_to_unicode(self, text):
        """
        Finds $...$ or $$...$$ blocks in 'text' and converts the inside to naive Unicode.
        """
        pattern = re.compile(r'(\${1,2})(.*?)(\1)', re.DOTALL)

        def replacer(m):
            eq_content = m.group(2).strip()
            return self.convert_latex_to_unicode(eq_content)

        return pattern.sub(replacer, text)

    def convert_latex_to_unicode(self, eq_text):
        """
        Naive approach: \alpha->α, x^2->x², x_2->x₂, etc.
        """
        # Common Greek letters
        greek_map = {
            r'\\alpha': 'α',
            r'\\beta': 'β',
            r'\\gamma': 'γ',
            r'\\delta': 'δ',
            r'\\theta': 'θ',
            r'\\mu': 'μ',
            r'\\pi': 'π',
            r'\\sigma': 'σ',
            r'\\phi': 'φ',
            r'\\omega': 'ω'
        }
        for latex_g, uni_g in greek_map.items():
            eq_text = eq_text.replace(latex_g, uni_g)

        # x^2 -> x²
        eq_text = re.sub(
            r'([A-Za-z0-9])\^([A-Za-z0-9])',
            lambda m: m.group(1) + self.to_superscript(m.group(2)),
            eq_text
        )
        # x_2 -> x₂
        eq_text = re.sub(
            r'([A-Za-z0-9])_([A-Za-z0-9])',
            lambda m: m.group(1) + self.to_subscript(m.group(2)),
            eq_text
        )

        # Common math symbols
        eq_text = eq_text.replace(r'\times', '×')
        eq_text = eq_text.replace(r'\cdot', '·')
        eq_text = eq_text.replace(r'\pm', '±')
        eq_text = eq_text.replace(r'\approx', '≈')
        eq_text = eq_text.replace(r'\neq', '≠')
        # optional spacing for '='
        eq_text = eq_text.replace(r'=', ' = ')

        return eq_text.strip()

    def to_superscript(self, char):
        supers = {
            '0': '⁰', '1': '¹', '2': '²', '3': '³',
            '4': '⁴', '5': '⁵', '6': '⁶', '7': '⁷',
            '8': '⁸', '9': '⁹',
            'n': 'ⁿ', 'i': 'ⁱ', '+': '⁺', '-': '⁻'
        }
        return supers.get(char, '^' + char)

    def to_subscript(self, char):
        subs = {
            '0': '₀', '1': '₁', '2': '₂', '3': '₃',
            '4': '₄', '5': '₅', '6': '₆', '7': '₇',
            '8': '₈', '9': '₉',
            '+': '₊', '-': '₋', '=': '₌', '(': '₍', ')': '₎'
        }
        return subs.get(char, '_' + char)

    def write_to_excel(self, mcq_data, excel_file):
        """Create a new Excel file and write MCQs data to it with the specified columns."""
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MCQs"

        # Define header based on the requested columns
        header = [
            "QuestionID",
            "Serial",
            "Class",
            "Subject",
            "Chapter",
            "Topic",
            "Question",
            "Ques_img",
            "OptionA",
            "OptionA_IMG",
            "OptionB",
            "OptionB_IMG",
            "OptionC",
            "OptionC_IMG",
            "OptionD",
            "OptionD_IMG",
            "Answer",
            "Explaination",
            "Explaination_IMG",
            "Hint",
            "Hint_img",
            "Difficulty_level",
            "Reference_Board/Institute",
            "Reference"
        ]

        # Write header row
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Get metadata values that will apply to all rows
        class_value = self.class_name.get().strip()
        subject_value = self.subject_name.get().strip()
        chapter_value = self.chapter_name.get().strip()

        # Write data rows - map the data to the new structure
        for row_num, row_data in enumerate(mcq_data, 2):
            # Debug information: print each row's data structure
            print(f"Row {row_num-1} data: Serial={row_data[0]}, Topic={row_data[3]}, " + 
                  f"Difficulty={row_data[4]}, Reference={row_data[5]}")
            
            # Data is now in format:
            # [0:serial, 1:question, 2:question_img, 3:topic, 4:difficulty, 5:board_inst, 
            #  6:option_a, 7:option_a_img, 8:option_b, 9:option_b_img, 10:option_c, 11:option_c_img, 
            #  12:option_d, 13:option_d_img, 14:answer, 15:explanation, 16:explanation_img, 
            #  17:hint, 18:hint_img]
            
            # Extract data
            serial_number = row_data[0]
            question_text = row_data[1]
            question_img = row_data[2]
            topic = row_data[3]
            difficulty = row_data[4]
            board_institute = row_data[5]
            option_a = row_data[6]
            option_a_img = row_data[7]
            option_b = row_data[8]
            option_b_img = row_data[9]
            option_c = row_data[10]
            option_c_img = row_data[11]
            option_d = row_data[12]
            option_d_img = row_data[13]
            answer = row_data[14]
            explanation = row_data[15]
            explanation_img = row_data[16]
            hint = row_data[17]
            hint_img = row_data[18]
            
            # For the answer column, ensure math expressions have $ delimiters
            # First standardize the answer format
            options_dict = {"ক": option_a, "খ": option_b, "গ": option_c, "ঘ": option_d}
            standardized_answer = self.standardize_answer(answer, options_dict)
            
            # Then ensure math equations have proper $ delimiters
            final_answer = self.ensure_equation_delimiters(standardized_answer)
            
            # Ensure LaTeX commands have proper backslash escaping
            final_answer = self.ensure_latex_escaped(final_answer)
            
            # Process other fields that might contain equations
            question_text = self.ensure_latex_escaped(question_text)
            option_a = self.ensure_latex_escaped(option_a)
            option_b = self.ensure_latex_escaped(option_b)
            option_c = self.ensure_latex_escaped(option_c)
            option_d = self.ensure_latex_escaped(option_d)
            explanation = self.ensure_latex_escaped(explanation)
            hint = self.ensure_latex_escaped(hint)
            
            # Finally preserve $ signs for Excel (only add ' where absolutely necessary)
            final_answer = self.preserve_dollar_signs(final_answer)
            question_text = self.preserve_dollar_signs(question_text)
            option_a = self.preserve_dollar_signs(option_a)
            option_b = self.preserve_dollar_signs(option_b)
            option_c = self.preserve_dollar_signs(option_c)
            option_d = self.preserve_dollar_signs(option_d)
            explanation = self.preserve_dollar_signs(explanation)
            hint = self.preserve_dollar_signs(hint)
            
            print(f"Original answer: {answer}")
            print(f"Standardized answer: {standardized_answer}")
            print(f"Final answer with delimiters and escaped LaTeX: {final_answer}")
            
            # Create row for Excel with all columns - use a dictionary for clarity
            # This approach makes it harder to accidentally extract code fragments
            excel_row = {
                "QuestionID": f"Q{serial_number}",
                "Serial": serial_number,
                "Class": class_value,
                "Subject": subject_value,
                "Chapter": chapter_value,
                "Topic": topic,
                "Question": question_text,
                "Ques_img": question_img,
                "OptionA": option_a,
                "OptionA_IMG": option_a_img,
                "OptionB": option_b,
                "OptionB_IMG": option_b_img,
                "OptionC": option_c,
                "OptionC_IMG": option_c_img,
                "OptionD": option_d,
                "OptionD_IMG": option_d_img,
                "Answer": final_answer,
                "Explaination": explanation, 
                "Explaination_IMG": explanation_img,
                "Hint": hint,
                "Hint_img": hint_img,
                "Difficulty_level": difficulty,
                "Reference_Board/Institute": board_institute,
                "Reference": ""                  # leave empty
            }
            
            # Convert dictionary to ordered list matching header
            new_row_data = [excel_row[col] for col in header]
            
            # Write the row to the worksheet
            for col_num, cell_value in enumerate(new_row_data, 1):
                # Set the cell value directly without modification
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Special handling for text cells that may contain equations ($ signs)
                if isinstance(cell_value, str) and '$' in cell_value:
                    # Don't add quotes - just set the number format to text
                    cell.number_format = '@'
                
                # Apply wrapping to text columns
                if col_num in [7, 9, 11, 13, 15, 18, 20]:  # Text columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-adjust column widths based on content (except for image columns)
        for col_idx, col in enumerate(ws.columns, 1):
            # Skip image columns for width adjustment
            if col_idx in [8, 10, 12, 14, 16, 19, 21]:
                continue
                
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column].width = adjusted_width
            
        # Set fixed width for image columns
        image_columns = [8, 10, 12, 14, 16, 19, 21]
        for col_idx in image_columns:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 30

        # Save the workbook
        wb.save(excel_file)
        
    def preserve_dollar_signs(self, text):
        """Ensure $ signs are preserved in Excel by using proper formatting"""
        if not text:
            return text
            
        # For Excel, only prefix with ' if it starts with a formula character
        # This prevents Excel from interpreting it as a formula
        if text.startswith('=') or text.startswith('+') or text.startswith('-'):
            text = "'" + text
        
        # We no longer need to add a ' prefix for cells with $ signs
        # Excel will display them correctly without it
            
        return text.strip()
        
    def ensure_equation_delimiters(self, text):
        """
        Ensure mathematical expressions have $ delimiters
        For answer column, we want to make sure equations are properly delimited
        """
        if not text:
            return text
            
        # If it's already an equation (has $ signs), just return it
        if '$' in text:
            return text
            
        # Common math symbols that indicate this is likely an equation
        math_symbols = ['+', '-', '=', '\\times', '\\cdot', '\\frac', '\\sqrt', '\\alpha', '\\beta', 
                       '\\sum', '\\int', '^', '_', '\\pi', '\\infty']
        
        # Check if the text contains any math symbols
        is_equation = any(symbol in text for symbol in math_symbols)
        
        # If it looks like an equation but doesn't have $ delimiters, add them
        if is_equation:
            # Make sure there are no double $ in case it's already partially delimited
            text = text.replace('$$', '$')
            # Add delimiters if not present
            if not text.startswith('$'):
                text = '$' + text
            if not text.endswith('$'):
                text = text + '$'
                
        return text

    def standardize_answer(self, answer, options):
        """
        Standardize the answer to A, B, C, D if it matches one of the options
        Otherwise, keep the original answer text
        
        Parameters:
            answer: The original answer text
            options: Dictionary of options {"ক": "option_a_text", "খ": "option_b_text", etc.}
        """
        # Clean the answer
        clean_answer = answer.strip()
        
        # Check if the answer is already A, B, C, D
        if clean_answer in ['A', 'B', 'C', 'D']:
            return clean_answer
            
        # Map Bengali option letters to English letters
        option_map = {"ক": "A", "খ": "B", "গ": "C", "ঘ": "D"}
        
        # Check if answer is a Bengali letter and convert
        if clean_answer in option_map:
            return option_map[clean_answer]
            
        # Try to find a match between the answer text and any option content
        for option_letter, option_text in options.items():
            if option_text and clean_answer == option_text.strip():
                return option_map.get(option_letter, option_letter)
                
        # If we get here, no match was found, return the original answer
        return clean_answer

    def add_tooltip(self, widget, text):
        """Add tooltip to a widget when mouse hovers over it"""
        def show_tooltip(event=None):
            x, y, _, _ = widget.bbox("insert")
            x += widget.winfo_rootx() + 25
            y += widget.winfo_rooty() + 25
            
            # Create a toplevel window
            self.tooltip = tk.Toplevel(widget)
            # Avoid window manager decorations
            self.tooltip.wm_overrideredirect(True)
            self.tooltip.wm_geometry(f"+{x}+{y}")
            
            label = tk.Label(self.tooltip, text=text, justify=tk.LEFT,
                            background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                            font=("Arial", "9", "normal"))
            label.pack(ipadx=5, ipady=5)
            
        def hide_tooltip(event=None):
            if hasattr(self, "tooltip"):
                self.tooltip.destroy()
                
        widget.bind("<Enter>", show_tooltip)
        widget.bind("<Leave>", hide_tooltip)

    def ensure_latex_escaped(self, text):
        """
        Ensure LaTeX commands have proper formatting
        For Excel display, LaTeX commands in $ delimited equations need single backslashes
        """
        if not text or '$' not in text:
            return text
            
        # Find all equations ($ delimited content)
        equation_pattern = re.compile(r'\$(.*?)\$')
        
        def format_latex_equation(match):
            eq_content = match.group(1)
            
            # Remove any \( and \) delimiters inside the equation
            eq_content = eq_content.replace('\\(', '').replace('\\)', '')
            
            # Common LaTeX commands that need a single backslash
            latex_commands = ['frac', 'sqrt', 'times', 'cdot', 'alpha', 'beta', 'gamma', 'delta', 'theta',
                             'sum', 'int', 'infty', 'pi', 'sin', 'cos', 'tan', 'log', 'ln', 'lim', 'text']
            
            # Add initial backslash if missing (e.g., if it starts with "frac" without a backslash)
            for cmd in latex_commands:
                # If it starts with the command name without a backslash, add one
                if eq_content.strip().startswith(cmd):
                    eq_content = '\\' + eq_content
                    break
                
                # Replace double backslashes with single backslashes for LaTeX commands
                eq_content = eq_content.replace('\\\\' + cmd, '\\' + cmd)
            
            # Remove extra spaces after backslashes (e.g., "\ frac" → "\frac")
            # Use simple string replacement instead of regex to avoid escape issues
            for cmd in latex_commands:
                eq_content = eq_content.replace('\\ ' + cmd, '\\' + cmd)  # One space
                eq_content = eq_content.replace('\\  ' + cmd, '\\' + cmd)  # Two spaces
                eq_content = eq_content.replace('\\   ' + cmd, '\\' + cmd)  # Three spaces
            
            # Ensure there's no extra space between $ and the equation content
            return '$' + eq_content.strip() + '$'
            
        # Apply formatting to each equation in the text
        return equation_pattern.sub(format_latex_equation, text)

    def extract_tables_from_html(self, html_file):
        """
        Extract tables from HTML file generated by pandoc
        Returns a list of HTML table elements as strings
        """
        with open(html_file, 'r', encoding='utf-8') as f:
            html_content = f.read()
            
        # Simple table extraction using regex
        table_pattern = re.compile(r'<table[^>]*>(.*?)</table>', re.DOTALL)
        tables = table_pattern.findall(html_content)
        
        # Save the number of tables found for reporting
        self.tables_found = len(tables)
        
        # Process each table to ensure it has proper HTML structure
        processed_tables = []
        for table_content in tables:
            # Ensure the table has proper HTML
            table_html = f"<table class='extracted-table'>{table_content}</table>"
            processed_tables.append(table_html)
            
        return processed_tables


def convert_docx_tables_to_html(docx_file, output_html=None):
    """
    Standalone function to extract tables from Word documents and convert them to HTML
    
    Args:
        docx_file: Path to the Word document
        output_html: Optional path for the HTML output file
        
    Returns:
        HTML string containing all tables from the document
    """
    import os
    import subprocess
    import re
    import tempfile
    
    # Create temp directory for processing
    with tempfile.TemporaryDirectory() as tmpdir:
        # Run pandoc to convert DOCX to HTML
        html_path = os.path.join(tmpdir, "temp.html")
        cmd = ["pandoc", docx_file, "-o", html_path]
        
        try:
            result = subprocess.run(cmd, check=True, capture_output=True, text=True)
            
            # Read the generated HTML
            with open(html_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
                
            # Extract tables using regex
            table_pattern = re.compile(r'<table[^>]*>(.*?)</table>', re.DOTALL)
            tables = table_pattern.findall(html_content)
            
            if not tables:
                print("No tables found in the document")
                return None
                
            # Create a complete HTML document with the tables
            output_html_content = """
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>Tables from Word Document</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    h1 { color: #333; }
                    table { border-collapse: collapse; width: 100%; margin-bottom: 30px; }
                    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                    th { background-color: #f2f2f2; font-weight: bold; }
                    tr:nth-child(even) { background-color: #f9f9f9; }
                    .table-container { margin-bottom: 40px; }
                </style>
            </head>
            <body>
                <h1>Tables Extracted from Word Document</h1>
            """
            
            for i, table_content in enumerate(tables):
                output_html_content += f"""
                <div class="table-container">
                    <h2>Table {i+1}</h2>
                    <table>{table_content}</table>
                </div>
                """
                
            output_html_content += """
            </body>
            </html>
            """
            
            # Save to output file if specified
            if output_html:
                with open(output_html, 'w', encoding='utf-8') as f:
                    f.write(output_html_content)
                print(f"Saved {len(tables)} tables to {output_html}")
                
            return output_html_content
            
        except subprocess.CalledProcessError as e:
            print(f"Error running pandoc: {e}")
            return None
        except Exception as e:
            print(f"Error processing document: {e}")
            return None


def main():
    root = tk.Tk()
    app = DocxToExcelPandocGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()